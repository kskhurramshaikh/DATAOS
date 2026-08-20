# NDI Assessment History -- Dev Queue item 5 ("NDI Assessment Dashboard
# + History", Master Directive Section 04).
#
# WHAT THIS IS: a durable, dated record of NDI assessments that were
# actually taken -- one row per recorded snapshot, storing the full
# 14-domain result exactly as it read at that moment, plus who recorded
# it and why. That is the "History" half of the item.
#
# WHAT THIS DELIBERATELY IS NOT: a generated trend line. The per-domain
# maturity/compliance INPUTS today are Dr. Saber's fixed BAJ demo
# baseline (banking_adapter.NDI_BAJ_BASELINE, his
# DataOS2_NDI_Methodology_Spec.pdf Section 2.6) -- an explicit
# instruction for this component, unlike IFRS 9 / SAMA / Customer 360
# which compute their inputs from real uploaded data. So two snapshots
# recorded a month apart WILL read identically unless those inputs
# change. Rather than manufacture a plausible improvement curve to make
# the page look alive, list_snapshots() below reports all_identical
# honestly and the History page says so on its face. Same standard
# already applied to Customer 360's clearly-labeled illustrative trend
# and to the NDI methodology note itself: a chart that implies real
# measurement where none exists is worse than an empty one.
#
# What the history genuinely records -- WHEN an assessment was taken,
# BY WHOM, and exactly what it said -- is real audit value on its own
# (it is the evidence trail a regulator asks for), and it becomes a
# real trend the moment the per-domain inputs become editable or
# data-derived. Nothing about this module changes when that happens.
#
# SCHEMA LOCATION NOTE, flagged rather than quietly done: this table's
# DDL lives here, not in app/db.py's init_db() alongside every other
# table in the codebase. It is a single self-contained feature table
# read and written only by this module, and _ensure_schema() below is
# idempotent on both backends. If a second module ever needs to read
# ndi_snapshots, move the DDL into db.py with the others rather than
# importing this one for its side effects.

import json

from app import db
from app.adapters import banking_adapter


def _ensure_schema():
    """CREATE TABLE IF NOT EXISTS on whichever backend is live.

    Called at the top of every public function rather than cached in a
    module-level "already created" flag on purpose: the test suite
    monkeypatches app.db.DB_PATH to a fresh SQLite file per test (see
    tests/test_dedup_adapter.py), so a cached flag would be stale the
    moment the underlying database changes underneath this process --
    exactly the class of silent-wrong-state bug the Postgres migration
    already cost several rounds to find."""
    # db._is_postgres() is module-private but this is the same package;
    # the alternative (db.storage_status()) opens a connection and runs
    # count queries just to answer a boolean.
    pk = "SERIAL PRIMARY KEY" if db._is_postgres() else "INTEGER PRIMARY KEY AUTOINCREMENT"
    with db.get_conn() as conn:
        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS ndi_snapshots (
                id {pk},
                recorded_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                recorded_by TEXT NOT NULL,
                note TEXT,
                display_score REAL NOT NULL,
                overall_maturity_score REAL NOT NULL,
                maturity_level TEXT NOT NULL,
                overall_compliance_pct REAL NOT NULL,
                overall_oe_score REAL NOT NULL,
                total_specs INTEGER NOT NULL,
                compliant_specs INTEGER NOT NULL,
                domains_json TEXT NOT NULL
            )
            """
        )
        conn.commit()


def _row_to_dict(row, include_domains: bool) -> dict:
    out = {
        "id": row["id"],
        "recorded_at": row["recorded_at"],
        "recorded_by": row["recorded_by"],
        "note": row["note"],
        "display_score": row["display_score"],
        "overall_maturity_score": row["overall_maturity_score"],
        "maturity_level": row["maturity_level"],
        "overall_compliance_pct": row["overall_compliance_pct"],
        "overall_oe_score": row["overall_oe_score"],
        "total_specs": row["total_specs"],
        "compliant_specs": row["compliant_specs"],
    }
    if include_domains:
        out["domains"] = json.loads(row["domains_json"])
    return out


def record_snapshot(recorded_by: str, note: str | None = None) -> dict:
    """Computes the current NDI assessment and stores it as a dated,
    attributed record. The full 14-domain breakdown is stored as it read
    at that moment (domains_json), NOT recomputed on read -- if the
    baseline or the methodology ever changes, an old snapshot must keep
    showing what was actually assessed then, or it isn't an audit
    record at all.

    recorded_by is required and not defaulted: an assessment record with
    nobody's name on it has no evidentiary value, and silently writing
    "system" or "anonymous" would make the History page look like a
    governance trail while being one in name only."""
    recorded_by = (recorded_by or "").strip()
    if not recorded_by:
        raise ValueError(
            "A name is required to record an assessment -- an audit record with no named "
            "recorder isn't an audit record."
        )
    note = (note or "").strip() or None

    assessment = banking_adapter.compute_ndi_assessment()

    _ensure_schema()
    with db.get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO ndi_snapshots (
                recorded_by, note, display_score, overall_maturity_score, maturity_level,
                overall_compliance_pct, overall_oe_score, total_specs, compliant_specs,
                domains_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                recorded_by,
                note,
                assessment["display_score"],
                assessment["overall_maturity_score"],
                assessment["maturity_level"],
                assessment["overall_compliance_pct"],
                assessment["overall_oe_score"],
                assessment["total_specs"],
                assessment["compliant_specs"],
                json.dumps(assessment["domains"]),
            ),
        )
        snapshot_id = cur.lastrowid
        conn.commit()

    return get_snapshot(snapshot_id)


def get_snapshot(snapshot_id: int) -> dict:
    """One recorded assessment in full, including the 14-domain
    breakdown exactly as stored at record time."""
    _ensure_schema()
    with db.get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM ndi_snapshots WHERE id = ?", (snapshot_id,)
        ).fetchone()
    if not row:
        raise ValueError(f"No recorded NDI assessment found with id {snapshot_id}.")
    return _row_to_dict(row, include_domains=True)


def list_snapshots(limit: int = 100) -> dict:
    """Every recorded assessment, newest first, each with its movement
    against the chronologically PREVIOUS record (not against the newest
    -- "what changed since last time" is the question a reviewer
    actually asks).

    Deltas are computed across the full ordered set before `limit` is
    applied, so the oldest row shown never gets a wrong delta just
    because the row before it fell outside the page.

    all_identical is the honest signal described in the module
    docstring: with the fixed BAJ baseline in place, every snapshot
    scores the same, and the History page renders that as a stated fact
    rather than a flat line the reader has to interpret."""
    _ensure_schema()
    with db.get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM ndi_snapshots ORDER BY recorded_at ASC, id ASC"
        ).fetchall()

    ordered = [_row_to_dict(r, include_domains=False) for r in rows]
    for i, entry in enumerate(ordered):
        if i == 0:
            entry["delta_display_score"] = None
            entry["delta_compliance_pct"] = None
        else:
            prev = ordered[i - 1]
            entry["delta_display_score"] = round(entry["display_score"] - prev["display_score"], 1)
            entry["delta_compliance_pct"] = round(
                entry["overall_compliance_pct"] - prev["overall_compliance_pct"], 1
            )

    newest_first = list(reversed(ordered))[:limit]
    distinct_scores = {e["display_score"] for e in ordered}

    return {
        "snapshots": newest_first,
        "count": len(ordered),
        "all_identical": len(ordered) > 1 and len(distinct_scores) == 1,
        "note": (
            "This is a record of assessments actually taken -- when, by whom, and exactly what "
            "each one said -- not a modelled trend. The per-domain maturity and compliance "
            "inputs are currently Dr. Saber's fixed BAJ demo baseline (his NDI methodology "
            "spec, Section 2.6), so every snapshot scores identically until those inputs "
            "become editable or data-derived. No synthetic movement has been added to make "
            "this look like real tracking."
        ),
    }


def compare_snapshots(id_a: int, id_b: int) -> dict:
    """Diffs any two recorded snapshots against each other -- period
    comparison, not just "vs the immediately previous record" the way
    list_snapshots()'s deltas work. Deliberately chronology-ordered
    regardless of which id the caller passed first (id_a/id_b are just
    "the two the user picked," not "from" and "to") -- the snapshot
    with the earlier recorded_at is always "from", the later one always
    "to", so a positive delta always means "improved since then" no
    matter which order they were selected in the UI.

    Domain deltas are matched by code, not position -- domains_json is
    stored per-snapshot exactly as computed at that time, and while the
    14-domain set hasn't changed since this was built, matching by code
    is the same defensive practice already used elsewhere in this
    codebase (see classification_adapter's keyword matching) rather
    than assuming two arrays stay aligned forever."""
    if id_a == id_b:
        raise ValueError("Pick two different assessments to compare.")

    snap_a = get_snapshot(id_a)
    snap_b = get_snapshot(id_b)

    # Order chronologically -- ties broken by id (matches list_snapshots'
    # own ORDER BY recorded_at ASC, id ASC).
    if (snap_a["recorded_at"], snap_a["id"]) <= (snap_b["recorded_at"], snap_b["id"]):
        earlier, later = snap_a, snap_b
    else:
        earlier, later = snap_b, snap_a

    domains_by_code_earlier = {d["code"]: d for d in earlier["domains"]}
    domains_by_code_later = {d["code"]: d for d in later["domains"]}
    all_codes = [d["code"] for d in later["domains"]]  # preserves later's own domain ordering

    domain_deltas = []
    for code in all_codes:
        d_from = domains_by_code_earlier.get(code)
        d_to = domains_by_code_later.get(code)
        if d_from is None or d_to is None:
            # Genuinely possible if the domain set itself ever changes
            # between the two recorded dates -- reported honestly
            # rather than silently skipped or defaulted to a fake 0.
            domain_deltas.append({
                "code": code,
                "name": d_to["name"] if d_to else d_from["name"],
                "comparable": False,
                "note": "This domain wasn't present in both snapshots.",
            })
            continue
        domain_deltas.append({
            "code": code,
            "name": d_to["name"],
            "comparable": True,
            "maturity_score_from": d_from["maturity_score"],
            "maturity_score_to": d_to["maturity_score"],
            "delta_maturity_score": round(d_to["maturity_score"] - d_from["maturity_score"], 2),
            "compliance_pct_from": d_from["compliance_pct"],
            "compliance_pct_to": d_to["compliance_pct"],
            "delta_compliance_pct": round(d_to["compliance_pct"] - d_from["compliance_pct"], 1),
        })

    return {
        "from": {
            "id": earlier["id"], "recorded_at": earlier["recorded_at"], "recorded_by": earlier["recorded_by"],
            "display_score": earlier["display_score"], "maturity_level": earlier["maturity_level"],
            "overall_compliance_pct": earlier["overall_compliance_pct"],
        },
        "to": {
            "id": later["id"], "recorded_at": later["recorded_at"], "recorded_by": later["recorded_by"],
            "display_score": later["display_score"], "maturity_level": later["maturity_level"],
            "overall_compliance_pct": later["overall_compliance_pct"],
        },
        "delta_display_score": round(later["display_score"] - earlier["display_score"], 1),
        "delta_compliance_pct": round(later["overall_compliance_pct"] - earlier["overall_compliance_pct"], 1),
        "delta_oe_score": round(later["overall_oe_score"] - earlier["overall_oe_score"], 3),
        "domains": domain_deltas,
        "identical": earlier["display_score"] == later["display_score"]
        and earlier["overall_compliance_pct"] == later["overall_compliance_pct"],
    }


def export_history_csv() -> str:
    """One row per recorded snapshot -- the same summary data
    list_snapshots() shows in the History table, as a real downloadable
    CSV for a compliance handoff. Domain-level detail isn't included
    here (14 columns x every snapshot would make this unreadable in a
    spreadsheet) -- see export_snapshot_csv() for that, per snapshot."""
    import csv
    import io

    data = list_snapshots(limit=100000)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "id", "recorded_at", "recorded_by", "note", "display_score", "maturity_level",
        "overall_compliance_pct", "overall_oe_score", "total_specs", "compliant_specs",
    ])
    # list_snapshots returns newest-first; export oldest-first, the
    # natural reading order for a record someone will open in Excel.
    for s in reversed(data["snapshots"]):
        writer.writerow([
            s["id"], s["recorded_at"], s["recorded_by"], s["note"] or "", s["display_score"],
            s["maturity_level"], s["overall_compliance_pct"], s["overall_oe_score"],
            s["total_specs"], s["compliant_specs"],
        ])
    return buf.getvalue()


def export_snapshot_csv(snapshot_id: int) -> str:
    """The full 14-domain breakdown for one recorded snapshot, as a
    real downloadable CSV -- everything DomainDetail shows on the
    History page, plus the fields the page doesn't render (spec_count,
    compliance_status, evidence) since a CSV handoff is exactly the
    place for the full record, not just the summary view."""
    import csv
    import io

    snap = get_snapshot(snapshot_id)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([f"NDI assessment #{snap['id']}", f"recorded {snap['recorded_at']}", f"by {snap['recorded_by']}"])
    if snap["note"]:
        writer.writerow([f"note: {snap['note']}"])
    writer.writerow([])
    writer.writerow([
        "code", "name", "spec_count", "maturity_score", "compliance_pct",
        "compliance_status", "is_oe_domain", "evidence",
    ])
    for d in snap["domains"]:
        writer.writerow([
            d["code"], d["name"], d["spec_count"], d["maturity_score"], d["compliance_pct"],
            d["compliance_status"], d["is_oe_domain"], d["evidence"],
        ])
    return buf.getvalue()


def export_history_xlsx() -> bytes:
    """Real .xlsx workbook of the full recorded history -- same rows as
    export_history_csv(), via openpyxl (already a project dependency,
    no new install needed) -- genuine formatted cells (bold header row,
    sized columns), not a CSV file renamed to .xlsx."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font

    data = list_snapshots(limit=100000)
    wb = Workbook()
    ws = wb.active
    ws.title = "NDI History"
    ws.append([
        "ID", "Recorded At", "Recorded By", "Note", "Display Score", "Maturity Level",
        "Compliance %", "OE Score", "Total Specs", "Compliant Specs",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for s in reversed(data["snapshots"]):
        ws.append([
            s["id"], s["recorded_at"], s["recorded_by"], s["note"] or "", s["display_score"],
            s["maturity_level"], s["overall_compliance_pct"], s["overall_oe_score"],
            s["total_specs"], s["compliant_specs"],
        ])
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_snapshot_xlsx(snapshot_id: int) -> bytes:
    """Real .xlsx of one recorded snapshot's full 14-domain breakdown --
    same fields export_snapshot_csv() writes."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font

    snap = get_snapshot(snapshot_id)
    wb = Workbook()
    ws = wb.active
    ws.title = f"NDI Snapshot {snap['id']}"
    ws.append([f"NDI assessment #{snap['id']}", f"recorded {snap['recorded_at']}", f"by {snap['recorded_by']}"])
    if snap["note"]:
        ws.append([f"note: {snap['note']}"])
    ws.append([])
    ws.append([
        "Code", "Name", "Spec Count", "Maturity Score", "Compliance %",
        "Compliance Status", "Is OE Domain", "Evidence",
    ])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for d in snap["domains"]:
        ws.append([
            d["code"], d["name"], d["spec_count"], d["maturity_score"], d["compliance_pct"],
            d["compliance_status"], d["is_oe_domain"], d["evidence"],
        ])
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_history_pdf() -> bytes:
    """Real PDF of the full recorded history, built with reportlab --
    pure Python, no system libraries required (unlike weasyprint/
    wkhtmltopdf, which need Cairo/Pango or a headless browser this
    image doesn't have). A genuine formatted table via reportlab's
    Table/TableStyle, not an HTML page screenshotted to PDF."""
    import io

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    data = list_snapshots(limit=100000)
    styles = getSampleStyleSheet()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), title="NDI Assessment History")
    elements = [Paragraph("NDI Assessment History", styles["Title"]), Spacer(1, 12)]
    if data["all_identical"]:
        elements.append(Paragraph(
            "Every recorded assessment scores identically -- expected while the per-domain "
            "inputs are the fixed BAJ demo baseline, not a bug.", styles["BodyText"],
        ))
        elements.append(Spacer(1, 10))

    rows = [["ID", "Recorded At", "Recorded By", "Note", "Display Score", "Maturity Level", "Compliance %"]]
    for s in reversed(data["snapshots"]):
        rows.append([
            str(s["id"]), s["recorded_at"], s["recorded_by"], s["note"] or "",
            str(s["display_score"]), s["maturity_level"], f'{s["overall_compliance_pct"]}%',
        ])
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F7A6B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D4D4D8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFB")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def export_snapshot_pdf(snapshot_id: int) -> bytes:
    """Real PDF of one recorded snapshot's full 14-domain breakdown."""
    import io

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    snap = get_snapshot(snapshot_id)
    styles = getSampleStyleSheet()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), title=f"NDI Assessment #{snap['id']}")
    elements = [
        Paragraph(f"NDI Assessment #{snap['id']}", styles["Title"]),
        Paragraph(f"Recorded {snap['recorded_at']} by {snap['recorded_by']}", styles["BodyText"]),
    ]
    if snap["note"]:
        elements.append(Paragraph(f"Note: {snap['note']}", styles["BodyText"]))
    elements.append(Spacer(1, 12))

    rows = [["Code", "Name", "Specs", "Maturity", "Compliance %", "Status", "OE Domain", "Evidence"]]
    for d in snap["domains"]:
        rows.append([
            d["code"], d["name"], str(d["spec_count"]), str(d["maturity_score"]),
            f'{d["compliance_pct"]}%', d["compliance_status"], "Yes" if d["is_oe_domain"] else "No",
            d["evidence"],
        ])
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F7A6B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D4D4D8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFB")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()
